from flask import Flask, request, jsonify, send_file
import os
import json
import csv
import io
from datetime import datetime
from flask_cors import CORS
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import logging
from logging.handlers import RotatingFileHandler
from werkzeug.middleware.proxy_fix import ProxyFix
from dotenv import load_dotenv

# Load environment variables
load_dotenv()

# Configure logging
logging.basicConfig(
    handlers=[RotatingFileHandler('app.log', maxBytes=100000, backupCount=3)],
    level=logging.INFO,
    format='[%(asctime)s] %(levelname)s in %(module)s: %(message)s'
)
logger = logging.getLogger(__name__)

app = Flask(__name__)
app.wsgi_app = ProxyFix(app.wsgi_app, x_proto=1, x_host=1)

# Configure CORS with specific options
CORS(app, resources={
    r"/api/*": {
        "origins": os.getenv('ALLOWED_ORIGINS', '*').split(','),
        "methods": ["GET", "POST", "DELETE"],
        "allow_headers": ["Content-Type"]
    }
})

# Security headers
@app.after_request
def add_security_headers(response):
    response.headers['X-Content-Type-Options'] = 'nosniff'
    response.headers['X-Frame-Options'] = 'SAMEORIGIN'
    response.headers['X-XSS-Protection'] = '1; mode=block'
    response.headers['Strict-Transport-Security'] = 'max-age=31536000; includeSubDomains'
    return response

# Directory to store project data
DATA_DIR = os.getenv('DATA_DIR', 'data')
if not os.path.exists(DATA_DIR):
    os.makedirs(DATA_DIR)

# Error handlers
@app.errorhandler(404)
def not_found_error(error):
    return jsonify({'error': 'Resource not found'}), 404

@app.errorhandler(500)
def internal_error(error):
    return jsonify({'error': 'Internal server error'}), 500

@app.route('/')
def index():
    try:
        return send_file('index.html')
    except Exception as e:
        logger.error(f"Error serving index.html: {str(e)}")
        return jsonify({'error': 'Error serving index file'}), 500

@app.route('/api/save-project', methods=['POST'])
def save_project():
    try:
        data = request.json
        
        # Validate required fields
        if not data or 'projectName' not in data:
            return jsonify({'error': 'Invalid data format'}), 400
        
        # Sanitize project name
        project_name = "".join(c for c in data['projectName'] if c.isalnum() or c in (' ', '-', '_')).strip()
        
        # Create a filename based on project name and timestamp
        filename = f"{project_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
        filepath = os.path.join(DATA_DIR, filename)
        
        # Add metadata
        data['savedAt'] = datetime.now().isoformat()
        
        # Save the project data as JSON
        with open(filepath, 'w') as f:
            json.dump(data, f, indent=2)
        
        logger.info(f"Project saved successfully: {filename}")
        return jsonify({
            'message': 'Project saved successfully',
            'filename': filename
        }), 200
        
    except Exception as e:
        logger.error(f"Error saving project: {str(e)}")
        return jsonify({'error': 'Error saving project'}), 500

@app.route('/api/projects', methods=['GET'])
def list_projects():
    try:
        projects = []
        
        for filename in os.listdir(DATA_DIR):
            if filename.endswith('.json'):
                filepath = os.path.join(DATA_DIR, filename)
                with open(filepath, 'r') as f:
                    data = json.load(f)
                    projects.append({
                        'filename': filename,
                        'projectName': data.get('projectName', 'Unnamed Project'),
                        'savedAt': data.get('savedAt', ''),
                        'columns': len(data.get('columnNames', [])),
                        'rows': len(data.get('rows', []))
                    })
        
        return jsonify(projects), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/projects/<filename>', methods=['GET'])
def get_project(filename):
    try:
        filepath = os.path.join(DATA_DIR, filename)
        
        if not os.path.exists(filepath):
            return jsonify({'error': 'Project not found'}), 404
        
        with open(filepath, 'r') as f:
            data = json.load(f)
        
        return jsonify(data), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/export/<filename>', methods=['GET'])
def export_project(filename):
    try:
        filepath = os.path.join(DATA_DIR, filename)
        
        if not os.path.exists(filepath):
            return jsonify({'error': 'Project not found'}), 404
        
        with open(filepath, 'r') as f:
            data = json.load(f)
        
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        
        # Write header row
        header = ['Task Name'] + data.get('columnNames', [])
        for col, header_text in enumerate(header, 1):
            ws.cell(row=1, column=col, value=header_text)
        
        # Format time helper function
        def format_time(ms):
            minutes = ms // 60000
            seconds = (ms % 60000) // 1000
            milliseconds = (ms % 1000) // 10
            return f"{minutes:02d}:{seconds:02d}.{milliseconds:02d}"
        
        # Write data rows
        for row_idx, row in enumerate(data.get('rows', []), 2):
            ws.cell(row=row_idx, column=1, value=row.get('name', 'Unnamed Task'))
            
            for col_idx in range(len(data.get('columnNames', []))):
                timer_id = f"{row.get('id', '')}-{col_idx}"
                timer_data = data.get('timerData', {}).get(timer_id, {})
                time_ms = timer_data.get('time', 0)
                ws.cell(row=row_idx, column=col_idx + 2, value=format_time(time_ms))
        
        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width
        
        # Save to BytesIO object
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f"{data.get('projectName', 'export')}.xlsx"
        )
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/delete/<filename>', methods=['DELETE'])
def delete_project(filename):
    try:
        filepath = os.path.join(DATA_DIR, filename)
        
        if not os.path.exists(filepath):
            return jsonify({'error': 'Project not found'}), 404
        
        os.remove(filepath)
        
        return jsonify({'message': 'Project deleted successfully'}), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    port = int(os.getenv('PORT', 5000))
    debug = os.getenv('FLASK_ENV') == 'development'
    host = os.getenv('HOST', '0.0.0.0')
    
    logger.info(f"Starting server on {host}:{port}")
    app.run(host=host, port=port, debug=debug)
